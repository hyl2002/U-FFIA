import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import torch.optim as optim
import torch
from sklearn import metrics
from sklearn.metrics import ConfusionMatrixDisplay
import torch.nn as nn
from utils.losses import get_loss_func
from torch.utils.tensorboard import SummaryWriter
import os
import time
import logging as log_config
import numpy as np
from utils.evaluate import Evaluator
from models.model_zoo.S3D import load_S3D_weight
import argparse
from utils.early_stopping import save_model
from tqdm import tqdm
from utils.early_stopping import save_model
from openpyxl import Workbook   # ★ 新增

def trainer(model, optimizer, train_loader, val_loader, test_loader, max_epoch, device, ckpt_dir, image_dir, scheduler):
    logger = log_config.getLogger()
    logger.info("Starting new training run")
    # ========================== 【★ 修改点 1】 ==========================
    # 原代码：loss_func = get_loss_func('clip_ce')
    # 新代码：使用 PyTorch 自带的 CrossEntropyLoss 并开启 label_smoothing
    # label_smoothing=0.1 意味着正确标签的目标概率是 0.9，而不是 1.0
    criterion = nn.CrossEntropyLoss(label_smoothing=0.1)
    # ================================================================
    loss_func = get_loss_func('clip_ce')
    evaluator = Evaluator(model=model)
    best_acc = 0
    best_epoch = 0
    last_5_epochs = []  # 保存最后5轮模型路径

    # --------------- Excel 初始化 ---------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Metrics"
    ws.append(["epoch", "train_loss", "val_acc", "val_mAP", "test_acc", "test_mAP","val_recall","val_f1"])  # 表头

    for epoch in range(max_epoch):
        debug_print = (epoch == 0)  # 只在第一个 epoch 打印一次
        mean_loss = 0
        # ★ 每个 epoch 开头打印一次学习率
        current_lr = optimizer.param_groups[0]['lr']
        print(f"\n==== Epoch {epoch} | Learning Rate = {current_lr} ====\n")
        for data_dict in tqdm(train_loader):
            data_dict['video_form'] = data_dict['video_form'].to(device)
            data_dict['waveform'] = data_dict['waveform'].to(device)
            data_dict['target'] = data_dict['target'].to(device)
            model.train()
            output_dict1 = model(data_dict['waveform'], data_dict['video_form'])
            # target_dict = {'target': data_dict['target']}
            # # ---------------- Debug 信息，只打印一次 ----------------
            # if debug_print:
            #     print("\n================ Debug Info ================")
            #     print("Target shape:", data_dict['target'].shape)
            #     print("Target example:", data_dict['target'])
            #     print("--------------------------------------------")
            #     print("Output shape:", output_dict1['clipwise_output'].shape)
            #     print("Output example (logits):", output_dict1['clipwise_output'][0])
            #     print("--------------------------------------------")
            # loss = loss_func(output_dict1, target_dict)
            # if debug_print:
            #     print("Loss value:", loss.item())
            #     print("============================================\n")
            #     debug_print = False

            # # --------------------------------------------------------
                      # ========================== 【★ 修改点 2】 ==========================
            # 原代码：loss = loss_func(output_dict1, target_dict)
            # 新代码：手动提取 logits 和 target 进行计算
            # output_dict1['clipwise_output'] 是模型的预测输出
            # data_dict['target'] 是真实标签
            
            logits = output_dict1['clipwise_output']
            targets = data_dict['target']
            loss = criterion(logits, targets)
            # ================================================================
            loss.backward()
            torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0)
            optimizer.step()
            scheduler.step()  # ★ 新增：更新学习率
            optimizer.zero_grad()
            mean_loss += loss.item()

        epoch_loss = mean_loss / len(train_loader)
        logger.info(f"Training loss {epoch_loss} at epoch {epoch}")



        if epoch % 1 == 0:
            model.eval()
            val_statistics = evaluator.evaluate_av(val_loader)
            val_cm = val_statistics['confu_matrix']
            val_mAP = np.mean(val_statistics['average_precision'])
            val_acc = np.mean(val_statistics['accuracy'])
            val_recall = val_statistics['recall']  # 新增
            val_f1 = val_statistics['f1']          # 新增
            val_message = val_statistics['message']

            if val_acc > best_acc:
                best_epoch = epoch
                best_acc = val_acc
                best_mAP = val_mAP
                best_cm = val_cm
                save_model(os.path.join(ckpt_dir, 'atten8_best.pt'), model, optimizer, val_acc, best_epoch)

            # 保存最后 5 轮模型
            if epoch >= max_epoch - 5:
                path = os.path.join(ckpt_dir, f'atten8_epoch{epoch}.pt')
                save_model(path, model, optimizer, val_acc, epoch)
                last_5_epochs.append((epoch, path))

            model.train()

            # --------------- 写入 Excel（测试值暂空） ---------------
            ws.append([epoch, epoch_loss, val_acc, val_mAP, None, None, val_recall, val_f1])

        logger.info(f'val_best_acc: {best_acc}, best mAP:{best_mAP}, best_epoch: {best_epoch}, current_val_acc: {val_acc},f1: {val_f1},recall: {val_recall}')


    # ------------------------- 测试 best model -------------------------
    logger.info('Evaluate on the Test dataset')
    model_path = os.path.join(ckpt_dir, 'atten8_best.pt')
    model.load_state_dict(torch.load(model_path)['model_state_dict'])
    model.eval()

    test_statistics = evaluator.evaluate_av(test_loader)
    test_cm = test_statistics['confu_matrix']
    ConfusionMatrixDisplay(test_cm).plot()
    plt.title("confusion_matrix_test")
    plt.savefig(os.path.join(image_dir, 'test-{}.png'.format(epoch)))
    plt.close()
    ave_precision = np.mean(test_statistics['average_precision'])
    ave_acc = np.mean(test_statistics['accuracy'])
    logger.info(f' test_dataset mAP: {ave_precision}, accuracy: {ave_acc}')

    # --------- 将 test 成绩写回对应 epoch 的行 ---------
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[0].value == best_epoch:
            row[4].value = ave_acc
            row[5].value = ave_precision
            break

    # ------------------------- 测试最后 5 轮模型 -------------------------
    last_5_acc = []
    last_5_mAP = []

    for ep, path in last_5_epochs:
        model.load_state_dict(torch.load(path)['model_state_dict'])
        model.eval()
        stats = evaluator.evaluate_av(test_loader)
        acc = np.mean(stats['accuracy'])
        mAP = np.mean(stats['average_precision'])
        last_5_acc.append(acc)
        last_5_mAP.append(mAP)

        logger.info(f'Epoch {ep} model test mAP: {mAP}, accuracy: {acc}')

        cm = stats['confu_matrix']
        ConfusionMatrixDisplay(cm).plot()
        plt.title(f"confusion_matrix_epoch{ep}")
        plt.savefig(os.path.join(image_dir, f'test_epoch{ep}.png'))

    all_acc = [ave_acc] + last_5_acc
    all_mAP = [ave_precision] + last_5_mAP
    avg_acc = np.mean(all_acc)
    avg_mAP = np.mean(all_mAP)
    logger.info(f'Average test accuracy (best + last 5 epochs): {avg_acc}, Average mAP: {avg_mAP}')

    # ------------------------- 保存 Excel -------------------------
    excel_path = os.path.join(ckpt_dir, "training_metrics.xlsx")
    wb.save(excel_path)
    logger.info(f"Metrics Excel saved to {excel_path}")
